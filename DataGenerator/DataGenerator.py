import openpyxl

def dataGenerator():
    file = openpyxl.load_workbook(r"D:\excelfile.xlsx")
    sheet = file["Arkusz1"]
    rows = sheet.max_row
    li_external = []
    li_internal = []
    for i in range(1, rows+1):
        li_internal=[]
        un = sheet.cell(i,1)
        up = sheet.cell(i,2)
        li_internal.insert(0, un.value)
        li_internal.insert(1, up.value)
        li_external.insert(i-1, li_internal)
    print(li_external)
    return li_external